import openpyxl
import os
path = 'E:\python\excel'
os.chdir(path)
print(os.getcwd())


wb = openpyxl.load_workbook('example.xlsx') # 接受文件名字符串，返回一个workbook数据类型的值，wb岱庙了这个excel文件
# 从工作簿中获取工作表 .sheetnames
print(wb.sheetnames)
